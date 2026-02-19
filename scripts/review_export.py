"""Export ingested documents + LLM-extracted metadata to XLSX for human review.

Usage:
    # Export for review
    python scripts/review_export.py export --db document_db.sqlite --output review_metadata.xlsx

    # Import corrections back
    python scripts/review_export.py import --db document_db.sqlite --input review_metadata.xlsx
"""
import sys
import os
import json
import argparse

# Add project root to path
script_dir = os.path.dirname(os.path.abspath(__file__))
project_root = os.path.dirname(script_dir)
if project_root not in sys.path:
    sys.path.insert(0, project_root)

from loguru import logger

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl not installed. Run: uv pip install openpyxl")


def export_for_review(db_path: str, output_path: str):
    """
    Export all documents and their LLM-extracted metadata to an XLSX spreadsheet.

    The spreadsheet includes blank columns for reviewer approval and notes.

    Args:
        db_path: Path to the SQLite database
        output_path: Path for the output .xlsx file
    """
    if not OPENPYXL_AVAILABLE:
        logger.error("openpyxl is required for export. Install with: uv pip install openpyxl")
        return

    from src.document_db import DocumentDatabase

    db = DocumentDatabase(db_path=db_path)
    documents = db.get_all_documents()
    db.close()

    if not documents:
        logger.warning("No documents found in database")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Metadata Review"

    # Header row
    headers = [
        "filename",
        "source_org",
        "procedure_category",
        "procedure_type",
        "doc_type",
        "age_group",
        "target_audience",
        "first_200_chars",
        "reviewer_approved",
        "reviewer_notes",
        "document_id",
    ]
    header_fill = PatternFill(start_color="2B5797", end_color="2B5797", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, doc in enumerate(documents, 2):
        content_preview = doc.get('content', '')[:200].replace('\n', ' ').strip()
        values = [
            doc.get('filename', ''),
            doc.get('source_org', ''),
            doc.get('procedure_category', ''),
            doc.get('procedure_type', ''),
            doc.get('doc_type', ''),
            doc.get('age_group', ''),
            doc.get('target_audience', ''),
            content_preview,
            '',  # reviewer_approved (blank for reviewer)
            '',  # reviewer_notes (blank for reviewer)
            doc.get('document_id', ''),
        ]
        for col, value in enumerate(values, 1):
            ws.cell(row=row_idx, column=col, value=value)

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    # Highlight review columns
    review_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    for row in ws.iter_rows(min_row=2, min_col=9, max_col=10):
        for cell in row:
            cell.fill = review_fill

    wb.save(output_path)
    logger.info(f"Exported {len(documents)} documents to {output_path}")
    print(f"✅ Exported {len(documents)} documents to {output_path}")


def import_reviewed(xlsx_path: str, db_path: str):
    """
    Read back a reviewed spreadsheet and update document metadata in the database.

    Args:
        xlsx_path: Path to the reviewed .xlsx file
        db_path: Path to the SQLite database
    """
    if not OPENPYXL_AVAILABLE:
        logger.error("openpyxl is required for import. Install with: uv pip install openpyxl")
        return

    from src.document_db import DocumentDatabase
    import sqlite3

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    # Read headers
    headers = [cell.value for cell in ws[1]]

    updated = 0
    db = DocumentDatabase(db_path=db_path)
    cursor = db.connection.cursor()

    for row in ws.iter_rows(min_row=2, values_only=False):
        row_data = {headers[i]: cell.value for i, cell in enumerate(row) if i < len(headers)}

        doc_id = row_data.get('document_id')
        if not doc_id:
            continue

        # Update metadata fields from the reviewed spreadsheet
        try:
            cursor.execute("""
                UPDATE documents SET
                    procedure_type = ?,
                    doc_type = ?,
                    age_group = ?,
                    target_audience = ?,
                    updated_at = datetime('now')
                WHERE document_id = ?
            """, (
                row_data.get('procedure_type', ''),
                row_data.get('doc_type', ''),
                row_data.get('age_group', ''),
                row_data.get('target_audience', ''),
                doc_id,
            ))
            updated += 1
        except Exception as e:
            logger.error(f"Error updating {doc_id}: {e}")

    db.connection.commit()
    db.close()

    logger.info(f"Updated {updated} documents from reviewed spreadsheet")
    print(f"✅ Updated {updated} documents from {xlsx_path}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Export/import metadata for human review")
    subparsers = parser.add_subparsers(dest="command", required=True)

    # Export command
    export_parser = subparsers.add_parser("export", help="Export metadata to XLSX")
    export_parser.add_argument("--db", type=str, default="document_db.sqlite",
                              help="Path to SQLite database")
    export_parser.add_argument("--output", type=str, default="review_metadata.xlsx",
                              help="Output XLSX file path")

    # Import command
    import_parser = subparsers.add_parser("import", help="Import reviewed XLSX back")
    import_parser.add_argument("--db", type=str, default="document_db.sqlite",
                              help="Path to SQLite database")
    import_parser.add_argument("--input", type=str, required=True,
                              help="Reviewed XLSX file path")

    args = parser.parse_args()

    if args.command == "export":
        export_for_review(args.db, args.output)
    elif args.command == "import":
        import_reviewed(args.input, args.db)
